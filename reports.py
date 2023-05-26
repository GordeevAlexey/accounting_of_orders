from io import BytesIO
from database.pg_db import BaseDB
from database.data import Period
from datetime import datetime, timedelta
from database.utils import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from send_email import Email
from typing import Any, Callable, Optional
import logging
from logger.logger import *



logger = logging.getLogger(__name__)


SUBORDERS_HEADER = (
    "id_orders",
    "content",
    "deadline",
    "status_code",
    "employee",
    "comment",
    "close_date"
)

ORDERS_HEADER = (
    "id",
    "issue_type",
    "issue_idx",
    "approving_date",
    "title",
    "initiator",
    "approving_employee",
    "employee",
    "deadline",
    "status_code",
    "comment"
)

REPORT_HEADER = (
    'Вид поручения',
    '№',
    'Дата утверждения',
    'Тема',
    'Инициатор',
    'Утверждающий руководитель',
    'Ответственные исполнители',
    'Срок исполнения',
    'Содержание поручения',
    'Статус поручения',
    'Примечание',
)

REPORT_HEADER_WIDTH = (17, 13, 23, 43, 40, 40, 40, 18, 45, 40, 18, 35)


def data_to_excel(ws, _fill_row: Callable,  data: list[tuple[Any]]) -> None:
        _idx = 0
        _id = None
        start = 7 if _fill_row.__qualname__.startswith("Approved") else 4

        for row_idx, tup in enumerate(data, start):
            if _idx != 0:
                row_idx = _idx + 1
            if _id is None or tup[0] != _id:
                _id = tup[0]
                ws[f'A{row_idx}'] = tup[1]
                ws[f'B{row_idx}'] = tup[2]
                ws[f'C{row_idx}'] = datetime.strftime(tup[3], "%d.%m.%Y")
                ws[f'D{row_idx}'] = tup[4]
                ws[f'E{row_idx}'] = tup[5]
                ws[f'F{row_idx}'] = tup[6]
                ws[f'G{row_idx}'] = tup[7]
                ws[f'H{row_idx}'] = datetime.strftime(tup[8], "%d.%m.%Y")
                ws[f'J{row_idx}'] = tup[9]
                ws[f'K{row_idx}'] = tup[10]
                row_idx += 1
            _fill_row(row_idx, tup[11:])
            ws[f'A{row_idx}'] = 'Поручение'
            ws[f'B{row_idx}'] = tup[2]
            ws[f'G{row_idx}'] = tup[15]
            ws[f'H{row_idx}'] = datetime.strftime(tup[13], "%d.%m.%Y")
            ws[f'I{row_idx}'] = tup[12]
            ws[f'J{row_idx}'] = tup[14]
            ws[f'K{row_idx}'] = tup[16]
            _idx = row_idx


class ExecutedOfThePeriod:

    def __init__(self, start_period: datetime, end_period: datetime, wb: Workbook) -> None:
        self.start_period = start_period
        self.end_period = end_period
        self.srp = datetime.strptime(start_period, "%Y-%m-%d").strftime("%d.%m.%Y")
        self.erp = datetime.strptime(end_period, "%Y-%m-%d").strftime("%d.%m.%Y")
        self.wb = wb
        self.wb.create_sheet("Исполненные за период")
        self.ws = wb['Исполненные за период']

    def _fill_row(self, row_idx: int, row: Any) -> None:
        if row[3] == "На исполнении":
            _fill = PatternFill(fill_type='solid', fgColor="E6B9B8")
            for col_idx in range(1, self.ws.max_column + 1):
                self.ws.cell(row_idx, col_idx).fill = _fill

    def _get_data(self):
        data = BaseDB().execute_query(
            f"""
            select 
            o.id,
            o.issue_type,
            o.issue_idx,
            o.approving_date,
            o.title,
            o.initiator,
            o.approving_employee,
            o.employee,
            o.deadline,
            o.status_code,
            o.comment,
            s.id_orders,
            s.content,
            s.deadline,
            s.status_code,
            s.employee,
            s.comment,
            s.close_date
            from orders as o,
            suborders as s
            where s.deadline >= '{self.start_period}'
            and s.deadline <= '{self.end_period}'
            and s.id_orders=o.id and s.deleted is false order by o.issue_idx 
            """
        )
        return data

    def _data_to_sheet(self) -> None:
        self.ws[
            "A1"] = f"Внутренние распорядительные документы со сроками исполнения в период с {self.srp} по {self.erp}"
        self.ws['A2'] = None

        self.ws.append(REPORT_HEADER)
        data = self._get_data()
        data_to_excel(self.ws, self._fill_row, data)

    def _apply_styles(self) -> None:
        self._data_to_sheet()
        self.ws.merge_cells('A1:K1')
        self.ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.ws['A1'].font = Font(name='Times New Roman', size=12, bold=True)
        self.ws.auto_filter.ref = "A3:K3"
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # HEADER
        for cell in self.ws['3:3']:
            cell.fill = PatternFill(fill_type='solid', fgColor="DBE5F1")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name='Times New Roman', size=12, bold=True)
            cell.border = border

        for row in self.ws.iter_rows(min_row=4):
            for col_idx, cell in enumerate(row, 1):
                self.ws.column_dimensions[get_column_letter(col_idx)].width = REPORT_HEADER_WIDTH[col_idx - 1]
                cell.font = Font(name='Times New Roman', size=12, bold=False)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = border

    def form(self) -> Workbook:
        self._apply_styles()
        return self.wb


class ApprovedForThePeriod:
    """
    Отчет об утвержденных за период ВРД
    """

    def __init__(self, start_period: datetime, end_period: datetime) -> None:
        self.start_period = start_period
        self.end_period = end_period
        self.srp = datetime.strptime(start_period, "%Y-%m-%d").strftime("%d.%m.%Y")
        self.erp = datetime.strptime(end_period, "%Y-%m-%d").strftime("%d.%m.%Y")
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Утвержденные за период"

    def _fill_row(self, row_idx: int, row: Any) -> None:
        if row[3] == "На исполнении":
            _fill = PatternFill(fill_type='solid', fgColor="E6B9B8")
        elif row[3] == "Заершено":
            _fill = PatternFill(fill_type='solid', fgColor="8DB4E3")
        elif row[2] == datetime.now().date().replace(month=12, day=31).strftime("%d.%m.%Y"):
            _fill = PatternFill(fill_type='solid', fgColor="D7E4BC")
        else:
            _fill = PatternFill(fill_type='solid', fgColor="FFFFFF")
        for col_idx in range(1, self.ws.max_column + 1):
            self.ws.cell(row_idx, col_idx).fill = _fill

    def _get_data(self):
        data = BaseDB().execute_query(
            f"""
            select 
            o.id,
            o.issue_type,
            o.issue_idx,
            o.approving_date,
            o.title,
            o.initiator,
            o.approving_employee,
            o.employee,
            o.deadline,
            o.status_code,
            o.comment,
            s.id_orders,
            s.content,
            s.deadline,
            s.status_code,
            s.employee,
            s.comment,
            s.close_date
            from orders as o,
            suborders as s
            where o.approving_date >= '{self.start_period}'
            and o.approving_date <= '{self.end_period}'
            and s.id_orders=o.id and s.deleted is false order by o.issue_idx 
            """
        )
        return data

    def _data_to_sheet(self) -> None:
        self.ws["A1"] = f"Внутренние распорядительные документы, утвержденные в период с {self.srp} по {self.erp}"
        self.ws["B2"] = '- без установленных сроков'
        self.ws["B3"] = '- на исполнении'
        self.ws["B4"] = '- завершено'
        self.ws['B5'] = None
        self.ws.append(REPORT_HEADER)
        data = self._get_data()
        data_to_excel(self.ws, self._fill_row, data)

    def _apply_styles(self):
        self.ws.merge_cells('A1:K1')
        self.ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.ws['A1'].font = Font(name='Times New Roman', size=12, bold=True)
        self.ws.auto_filter.ref = "A6:K6"
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Ячейки с расшифровкой заливки
        for cell in ('A2', 'A3', 'A4'):
            self.ws[cell].border = border
        self.ws['A2'].fill = PatternFill(fill_type='solid', fgColor="D7E4BC")
        self.ws['A3'].fill = PatternFill(fill_type='solid', fgColor="E6B9B8")
        self.ws['A4'].fill = PatternFill(fill_type='solid', fgColor="8DB4E3")

        # HEADER
        for cell in self.ws['6:6']:
            cell.fill = PatternFill(fill_type='solid', fgColor="DBE5F1")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name='Times New Roman', size=12, bold=True)
            cell.border = border

        for row in self.ws.iter_rows(min_row=7):
            for col_idx, cell in enumerate(row, 1):
                self.ws.column_dimensions[get_column_letter(col_idx)].width = REPORT_HEADER_WIDTH[col_idx - 1]
                cell.font = Font(name='Times New Roman', size=12, bold=False)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = border

    def form(self) -> Workbook:
        self._data_to_sheet()
        self._apply_styles()
        logger.info(f'Отчет об утвержденных за период {self.srp}-{self.erp} ВРД')
        return self.wb


class WeeklyReport:
    def __init__(self) -> None:
        self.wb = None

    def report_time(self) -> Optional[bool]:
        self.report_date = datetime.now()
        if self.report_date.weekday() == 4:
            logger.info(f'Проверка текущего времени для отчета -> {self.report_date}')
            self.time_to_report = True
            self.start_report_period = self.report_date - timedelta(days=11)
            self.end_report_period = self.start_report_period + timedelta(days=6)
            self.start_report_period = self.start_report_period.strftime("%Y-%m-%d")
            self.end_report_period = self.end_report_period.strftime("%Y-%m-%d")
            self.output = BytesIO()
            return True

    def form_approved_for_the_period(self) -> None:
        approved_period = ApprovedForThePeriod(
            self.start_report_period,
            self.end_report_period,
            ).form()
        self.wb = approved_period

    def form_executed_for_the_period(self) -> bytes:
        wb = ExecutedOfThePeriod(
            self.start_report_period,
            self.end_report_period,
            self.wb
        ).form()
        wb.save(self.output)
        res = self.output.getvalue()
        self.output.close()
        self.output = BytesIO()
        return res

    def send_report(self) -> None:
        if self.report_time() is True:
            self.form_approved_for_the_period()
            srp = '.'.join(self.start_report_period.split("-")[::-1])
            erp = '.'.join(self.end_report_period.split("-")[::-1])
            Email.send_weekly_report(
                f"""Отчет об исполнении за период {srp} - {erp}\n\n"""
                """*Данное сообщение сформировано автоматически. Не нужно на него отвечать.\n\n""",
                f"""weekly_report {srp}-{erp}""",
                self.form_executed_for_the_period()
            )
            logger.info(f'Отчет отправлен за период -> {srp}-{erp}.')


def manual_report_unloading(period: Period) -> bytes:
    """
    Ручная выгрузка отчета
    """
    with BytesIO() as output:
        wb = ExecutedOfThePeriod(**period.dict(), wb=Workbook()).form()
        del wb['Sheet']
        wb.save(output)
        res = output.getvalue()
    return res 
